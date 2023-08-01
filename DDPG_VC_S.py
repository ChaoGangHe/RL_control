# -*- coding: utf-8 -*-
"""
Created on Mon Dec 13 10:41:53 2021

@author: admin

Implementation of Deep Deterministic Policy Gradients (DDPG) with pytorch 
riginal paper: https://arxiv.org/abs/1509.02971
Not the author's implementation !

"""
import argparse
import copy
import math

import os
import numpy as np
import pandas as pd

import torch
import torch.nn as nn
import torch.nn.functional as F
import torch.optim as optim

from tensorboardX import SummaryWriter
from scipy.fftpack import fft
import matplotlib.pyplot as plt
import openpyxl
#%% 超参数
parser = argparse.ArgumentParser()
parser.add_argument('--mode', default='train', type=str)         # mode = 'train' or 'test'
parser.add_argument('--load_nn', default=False, type=bool)       #是否导入已有网络

parser.add_argument('--tau',  default=0.005, type=float)         # target smoothing coefficient

parser.add_argument('--lr_A', default=3e-4, type=float)          #A网络学习率
parser.add_argument('--lr_C', default=3e-4, type=float)          #C网络学习率

parser.add_argument('--gamma', default=0.99, type=int)              # discounted factor
parser.add_argument('--capacity', default=2000, type=int)       # replay buffer size
parser.add_argument('--batch_size', default=128, type=int)        # mini batch size

parser.add_argument('--episode_length', default=300, type=int) # 回合长度()===x0的长度
parser.add_argument('--save_interval', default=1, type=int)      # 相隔n回合存储一次网络参数
parser.add_argument('--max_episode', default=2001, type=int)      # 回合数
parser.add_argument('--update_iteration', default=300, type=int) # 每回合更新网络参数的次数

parser.add_argument('--exploration_noise', default=0.2, type=float)   #探索噪声初值

parser.add_argument('--state_scale_factor', default=0.1, type=float)  #状态放大系数（放大之后输入网络） 10
parser.add_argument('--action_scale_factor', default=0.02, type=float)  #动作放大系数（网络输出后放大） 2
parser.add_argument('--test_num',default=300,type=int)#训练样本长度
args = parser.parse_args()

device = 'cuda' if torch.cuda.is_available() else 'cpu'
# device = torch.device('cpu')

#%%路径
path = os.getcwd()

path1 = path + '\\pic2\\'
if not os.path.exists(path1)==True:
    os.mkdir(path1)

directory = path + '\\nn2\\'
if not os.path.exists(directory)==True:
    os.mkdir(directory)

Q_value = path + '\\Q_value\\'
if not os.path.exists(Q_value)==True:
    os.mkdir(Q_value)

loss_value = path + '\\loss_value\\'
if not os.path.exists(loss_value)==True:
    os.mkdir(loss_value)

#%%读取传递通道模型参数
df = pd.read_excel('Wts0415.xlsx',sheet_name='Sheet1', header=None)
#df = pd.read_excel('Wts0126.xlsx',sheet_name='Sheet1', header=None)
data = df.values

W01 = data[:,0]
W02 = data[:,1]
W03 = data[:,2]
W04 = data[:,3]

W11 = data[:,4]
W12 = data[:,5]
W13 = data[:,6]
W14 = data[:,7]

W21 = data[:,8]
W22 = data[:,9]
W23 = data[:,10]
W24 = data[:,11]

W31 = data[:,12]
W32 = data[:,13]
W33 = data[:,14]
W34 = data[:,15]

W41 = data[:,16]
W42 = data[:,17]
W43 = data[:,18]
W44 = data[:,19]


W01r = W01[::-1]
W02r = W02[::-1]
W03r = W03[::-1]
W04r = W04[::-1]

W11r = W11[::-1]
W12r = W12[::-1]
W13r = W13[::-1]
W14r = W14[::-1]

W21r = W21[::-1]
W22r = W22[::-1]
W23r = W23[::-1]
W24r = W24[::-1]

W31r = W31[::-1]
W32r = W32[::-1]
W33r = W33[::-1]
W34r = W34[::-1]

W41r = W41[::-1]
W42r = W42[::-1]
W43r = W43[::-1]
W44r = W44[::-1]#(300,1)

#df1 = pd.read_excel('mortor_self.xlsx',sheet_name='motor', header=None)
df1 = pd.read_excel('motor0on0415_test4.xlsx',sheet_name='data', header=None)
#df1 = pd.read_excel('motor0on.xlsx',sheet_name='Data', header=None)
ii = 1
x0_array = df1.values[1+ii:20001+ii,0]

Q_data = openpyxl.Workbook()
sheet_score = Q_data.create_sheet('value',0)
sheet_score.append(['current','target'])
Q_data.save('Q_data.xlsx')

loss_data = openpyxl.Workbook()
sheet_loss = loss_data.create_sheet('value',0)
sheet_loss.append(['loss',])
loss_data.save('loss_data.xlsx')
#%% 定义环境
class Env():
    def __init__(self):
        #振源、电机的输入均是长度300的向量（）
        self.X0 = 300*[0]
        self.X1 = 300*[0]
        self.X2 = 300*[0]
        self.X3 = 300*[0]
        self.X4 = 300*[0]
        self.tau = 0.001
    
    def step(self, x0,x1,x2,x3,x4):
        #x0是振源, x1,x2,x3,x4电机当前输入值、也是Action
        #更新输入向量
        self.X0.append(x0)
        self.X0.pop(0)
        
        self.X1.append(x1)
        self.X1.pop(0)

        self.X2.append(x2)
        self.X2.pop(0)

        self.X3.append(x3)
        self.X3.pop(0)

        self.X4.append(x4)
        self.X4.pop(0)
        
        #更新传感器的输出
        self.y1 =  np.array(self.X0).dot(W01r) + \
                   np.array(self.X1).dot(W11r) + \
                   np.array(self.X2).dot(W21r) + \
                   np.array(self.X3).dot(W31r) + \
                   np.array(self.X4).dot(W41r)

        self.y2 =  np.array(self.X0).dot(W02r) + \
                   np.array(self.X1).dot(W12r) + \
                   np.array(self.X2).dot(W22r) + \
                   np.array(self.X3).dot(W32r) + \
                   np.array(self.X4).dot(W42r)
                 
        self.y3 =  np.array(self.X0).dot(W03r) + \
                   np.array(self.X1).dot(W13r) + \
                   np.array(self.X2).dot(W23r) + \
                   np.array(self.X3).dot(W33r) + \
                   np.array(self.X4).dot(W43r)
                   
        self.y4 =  np.array(self.X0).dot(W04r) + \
                   np.array(self.X1).dot(W14r) + \
                   np.array(self.X2).dot(W24r) + \
                   np.array(self.X3).dot(W34r) + \
                   np.array(self.X4).dot(W44r)



        # self.y1 = self.y1 *0.01
        # self.y2 = self.y2 *0.01
        # self.y3 = self.y3 *0.01
        # self.y4 = self.y4 *0.01
        self.state = np.array([self.y1, self.y2, self.y3, self.y4])

        self.done = 0
        
        # 定义reward
        self.reward = -100*(self.y1)**2 -100*(self.y2)**2 -100*(self.y3)**2 -100*(self.y4)**2
      
        return self.state, self.reward, self.done
    
    def reset(self):
        self.X0 = 300 * [0]
        self.X1 = 300 * [0]
        self.X2 = 300 * [0]
        self.X3 = 300 * [0]
        self.X4 = 300 * [0]
        self.tau = 0.001
        self.state = np.array([0.0, 0.0, 0.0, 0.0])
        return self.state

#%% 定义智能体
class Replay_buffer():
    '''
    Code based on:
    https://github.com/openai/baselines/blob/master/baselines/deepq/replay_buffer.py
    Expects tuples of (state, next_state, action, reward, done)
    '''
    def __init__(self, max_size=args.capacity):
        self.storage = []
        self.max_size = max_size
        self.ptr = 0

    def push(self, data):
        if len(self.storage) == self.max_size:
            self.storage[int(self.ptr)] = data
            self.ptr = (self.ptr + 1) % self.max_size
        else:
            self.storage.append(data)

    def sample(self, batch_size):
        ind = np.random.randint(0, len(self.storage), size=batch_size)
        x, y, u, r, d = [], [], [], [], []

        for i in ind:
            X, Y, U, R, D = self.storage[i]
            x.append(np.array(X, copy=False))
            y.append(np.array(Y, copy=False))
            u.append(np.array(U, copy=False))
            r.append(np.array(R, copy=False))
            d.append(np.array(D, copy=False))

        return np.array(x), np.array(y), np.array(u), np.array(r).reshape(-1, 1), np.array(d).reshape(-1, 1)



class Actor1(nn.Module):
    def __init__(self, state_dim, action_dim, max_action):
        super(Actor1, self).__init__()
        self.l1 = nn.Linear(state_dim, 256)
        self.l2 = nn.Linear(256, 256)
        self.l3 = nn.Linear(256, action_dim)

        self.max_action = max_action

    def forward(self, state):
        a = F.relu(self.l1(state))
        a = F.relu(self.l2(a))
        return self.max_action * torch.tanh(self.l3(a))

class Critic1(nn.Module):
    def __init__(self, state_dim, action_dim):
        super(Critic1, self).__init__()
        # Q1 architecture
        self.l1 = nn.Linear(state_dim + action_dim, 256)
        self.l2 = nn.Linear(256, 256)
        self.l3 = nn.Linear(256, 1)

        # Q2 architecture
        self.l4 = nn.Linear(state_dim + action_dim, 256)
        self.l5 = nn.Linear(256, 256)
        self.l6 = nn.Linear(256, 1)

    def forward(self, state, action):
        sa = torch.cat([state, action], 1)

        q1 = F.relu(self.l1(sa))
        q1 = F.relu(self.l2(q1))
        q1 = self.l3(q1)

        q2 = F.relu(self.l4(sa))
        q2 = F.relu(self.l5(q2))
        q2 = self.l6(q2)
        return q1, q2

    def Q1(self,state,action):
        sa = torch.cat([state, action], 1)

        q1 = F.relu(self.l1(sa))
        q1 = F.relu(self.l2(q1))
        q1 = self.l3(q1)
        return q1

class DDPG(object):
    def __init__(self, state_dim, action_dim, max_action):
        self.actor = Actor1(state_dim, action_dim, max_action).to(device)
        self.actor_target = copy.deepcopy(self.actor)
        self.actor_optimizer = optim.Adam(self.actor.parameters(), lr = args.lr_A)

        self.critic = Critic1(state_dim, action_dim).to(device)
        self.critic_target = copy.deepcopy(self.critic)
        self.critic_optimizer = optim.Adam(self.critic.parameters(), lr = args.lr_C)
        self.replay_buffer = Replay_buffer()
        self.writer = SummaryWriter(directory)

        self.num_critic_update_iteration = 0
        self.num_actor_update_iteration = 0
        self.num_training = 0

    def select_action(self, state):
        state = torch.FloatTensor(state.reshape(1, -1)).to(device)
        return self.actor(state).cpu().data.numpy().flatten()

    def update(self):

        for it in range(args.update_iteration):
            # Sample replay buffer
            x, y, u, r, d = self.replay_buffer.sample(args.batch_size)
            state = torch.FloatTensor(x).to(device)
            action = torch.FloatTensor(u).to(device)
            next_state = torch.FloatTensor(y).to(device)
            done = torch.FloatTensor(1-d).to(device)
            reward = torch.FloatTensor(r).to(device)

            # Select action according to policy and add clipped noise
            noise = (
                    torch.randn_like(action) * 0.2
            ).clamp(-0.5, 0.5)

            next_action = (
                    self.actor_target(next_state) + noise
            ).clamp(-self.actor.max_action, self.actor.max_action)

            # Compute the target Q value
            target_Q1, target_Q2 = self.critic_target(next_state, next_action)
            target_Q = torch.min(target_Q1, target_Q2)
            target_Q = reward + done * args.gamma * target_Q

            # Get current Q estimates
            current_Q1, current_Q2 = self.critic(state, action)

            # Compute critic loss
            critic_loss = F.mse_loss(current_Q1, target_Q) + F.mse_loss(current_Q2, target_Q)

            # Optimize the critic
            self.critic_optimizer.zero_grad()
            critic_loss.backward()
            self.critic_optimizer.step()

            # Delayed policy updates
            if it % 2 == 0:

                # Compute actor losse
                actor_loss = -self.critic.Q1(state, self.actor(state)).mean()

                # Optimize the actor
                self.actor_optimizer.zero_grad()
                actor_loss.backward()
                self.actor_optimizer.step()

                # Update the frozen target models
                for param, target_param in zip(self.critic.parameters(), self.critic_target.parameters()):
                    target_param.data.copy_(args.tau * param.data + (1 - args.tau) * target_param.data)

                for param, target_param in zip(self.actor.parameters(), self.actor_target.parameters()):
                    target_param.data.copy_(args.tau * param.data + (1 - args.tau) * target_param.data)



            self.num_actor_update_iteration += 1
            self.num_critic_update_iteration += 1

    def save(self):
        torch.save(self.actor.state_dict(), directory + 'actor.pt')
        torch.save(self.critic.state_dict(), directory + 'critic.pt')
        print("====================================")
        print("Model has been saved...")
        print("====================================")

    def load(self):
        self.actor.load_state_dict(torch.load(directory + 'actor.pt'))
        self.critic.load_state_dict(torch.load(directory + 'critic.pt'))
        print("====================================")
        print("model has been loaded...")
        print("====================================")


#%% 测试环境
agent = DDPG(state_dim=400, action_dim=4, max_action=1) 

'''
for actor_params in agent.actor.parameters():
    torch.nn.init.normal_(actor_params ,mean=0, std=0.1)
    #torch.nn.init.uniform_(actor_params,-1,1)
print("============ agent actor initialled ================")

for critic_params in agent.critic.parameters():
    torch.nn.init.normal_(critic_params, mean=0, std=0.1)
    #torch.nn.init.uniform_(critic_params,-1,1)
print("============ agent critic initialled ================")
'''

env = Env()

#测试env
j = 0
dt = env.tau

f1 = 35
A1 = 2
fai1 = np.pi*1/2

f2 = 40
A2 = 1
fai2 = np.pi*4/3

f3 = 45
A3 = 2   
fai3 = np.pi*5/3

acc1__ = []
acc2__ = []
acc3__ = []
acc4__ = []

S_acc_org = 0
while j <= args.episode_length-1:
    # 开启振源, 实际中的振源不可知
    # x0 = A * np.sin(2*np.pi*f * dt * j)
    x0 = x0_array[j]
                    
    #x0 = A1*np.sin(2*np.pi*f1*j*dt + fai1) + A2*np.sin(2*np.pi*f2*j*dt + + fai2) + A3*np.sin(2*np.pi*f3*j*dt + fai3)
     
    # 电机的动作，神经网络的输出，这里关闭电机
    x1 = 0
    x2 = 0
    x3 = 0
    x4 = 0
    # 状态，实际从传感器读取
    env.state, reward, done = env.step(x0,x1,x2,x3,x4)
    acc1__.append(env.state[0])
    acc2__.append(env.state[1])
    acc3__.append(env.state[2])
    acc4__.append(env.state[3])
    
    j += 1
S_acc_org1 = 0
S_acc_org2 = 0
S_acc_org3 = 0
S_acc_org4 = 0
for i in acc1__:
    S_acc_org1 += i * i
for i in acc2__:
    S_acc_org2 += i * i
for i in acc3__:
    S_acc_org3 += i * i
for i in acc4__:
    S_acc_org4 += i * i
S_acc_org1 = S_acc_org1/len(acc1__)
S_acc_org2 = S_acc_org2/len(acc1__)
S_acc_org3 = S_acc_org3/len(acc1__)
S_acc_org4 = S_acc_org4/len(acc1__)
S_acc_org = (S_acc_org1+S_acc_org2+S_acc_org3+S_acc_org4)/4
# print("acc1__:", acc1__)#红色
#
# print("acc2__:", acc2__)
#
# print("acc3__:", acc3__)
#
# print("acc4__:", acc4__)



Fs = int(1/env.tau)
nfft = 2*Fs
fre = Fs/nfft*(np.array(range(0,int(nfft/2))))

Y1 = fft(acc1__, nfft)
P1 = abs(Y1/nfft)

Y2 = fft(acc2__, nfft)
P2 = abs(Y2/nfft)

Y3 = fft(acc3__, nfft)
P3 = abs(Y3/nfft)

Y4 = fft(acc4__, nfft)
P4 = abs(Y4/nfft)

# plt.figure(figsize=(10,6))
# plt.subplot(221)
# plt.plot(fre, P1[0:int(nfft/2),], '-b')
# plt.xlabel('Frequency(Hz)')
# plt.ylabel('Acc1 FFT')
#
# plt.subplot(222)
# plt.plot(fre, P2[0:int(nfft/2),], '-b')
# plt.xlabel('Frequency(Hz)')
# plt.ylabel('Acc2 FFT')
#
# plt.subplot(223)
# plt.plot(fre, P3[0:int(nfft/2),], '-b')
# plt.xlabel('Frequency(Hz)')
# plt.ylabel('Acc3 FFT')
#
# plt.subplot(224)
# plt.plot(fre, P4[0:int(nfft/2),], '-b')
# plt.xlabel('Frequency(Hz)')
# plt.ylabel('Acc4 FFT')

t = env.tau * np.array(range(0, args.episode_length))
plt.figure(figsize=(16, 12))

plt.subplot(441)
plt.plot(t, acc1__, "r--", label="acc1", linewidth=3)
plt.xlabel("Time(s)")
plt.ylabel("Acc")
plt.legend()

plt.subplot(442)
plt.plot(t, acc2__, "r--", label="acc2", linewidth=3)
plt.ylabel("Acc")
plt.xlabel("Time(s)")
plt.legend()

plt.subplot(443)
plt.plot(t, acc3__, "r--", label="acc3", linewidth=3)
plt.ylabel("Acc")
plt.xlabel("Time(s)")
plt.legend()

plt.subplot(444)
plt.plot(t, acc4__, "r--", label="acc4", linewidth=3)
plt.ylabel("Acc")
plt.xlabel("Time(s)")
plt.legend()

plt.subplot(445)
plt.plot(fre, P1[0:int(nfft / 2), ], "r--", label="acc1", linewidth=3)
plt.xlabel('Frequency(Hz)')
plt.ylabel('FFT')
plt.legend()

plt.subplot(446)
plt.plot(fre, P2[0:int(nfft / 2), ], "r--", label="acc2", linewidth=3)
plt.xlabel('Frequency(Hz)')
plt.ylabel('FFT')
plt.legend()

plt.subplot(447)
plt.plot(fre, P3[0:int(nfft / 2), ], "r--", label="acc3", linewidth=3)
plt.xlabel('Frequency(Hz)')
plt.ylabel('FFT')
plt.legend()

plt.subplot(448)
plt.plot(fre, P4[0:int(nfft / 2), ], "r--", label="acc4", linewidth=3)
plt.xlabel('Frequency(Hz)')
plt.ylabel('FFT')
plt.legend()

# t = env.tau * np.array(range(0, args.episode_length))
# plt.figure(figsize=(16, 12))
# plt.subplot(441)
# plt.plot(t, acc1__, "r--", label="acc1", linewidth=3)
# plt.xlabel("Time(s)")
# plt.ylabel("Acc")
plt.savefig(path + '\\pic\\' +  'org.png')
#plt.show()

#%%训练智能体
if args.mode == 'train':
    print("============ train agent ==============")
    if args.load_nn == True:
        agent.load()
    S_acc_current = 0
    ep_reward__ = []
    for i_episode in range(args.max_episode):#2000
        # 进入新的episode之前, 清空列表
        acc1c__ = []
        acc2c__ = []
        acc3c__ = []
        acc4c__ = []
    
        action1__ = []
        action2__ = []
        action3__ = []
        action4__ = []
    
        ep_reward = 0
        num = 0 #(0,2000)
    
        # 重置
        env.reset()
    
        s1_history = 100*[0.0]
        s2_history = 100*[0.0]
        s3_history = 100*[0.0]
        s4_history = 100*[0.0]  
        s_history = s1_history + s2_history + s3_history + s4_history
    
        # 获取第一个s
        state = args.state_scale_factor *np.array(s_history)
        
        args.exploration_noise *= 0.999 #每回合噪声衰减
    
        while(1):
            action = agent.select_action(state)
            #action[2] = 0
            #action[3] = 0
            action = (action + np.random.normal(0, args.exploration_noise, size=4)).clip(-1, 1)
    
            #x0 = A*np.sin(2*np.pi*f*num*dt)
            x0 = x0_array[num]
            
            env.state, reward, done = env.step(x0, args.action_scale_factor*action[0], args.action_scale_factor*action[1], args.action_scale_factor*action[2], args.action_scale_factor*action[3])

            s1_history = [i * 0.5 for i in s1_history]
            s1_history.append(env.state[0])
            s1_history.pop(0)
            s2_history = [i * 0.5 for i in s2_history]
            s2_history.append(env.state[1])
            s2_history.pop(0)
            s3_history = [i * 0.5 for i in s3_history]
            s3_history.append(env.state[2])
            s3_history.pop(0)
            s4_history = [i * 0.5 for i in s4_history]
            s4_history.append(env.state[3])
            s4_history.pop(0)
            
            s_history = s1_history + s2_history + s3_history + s4_history
            
            next_state = args.state_scale_factor * np.array(s_history)
            
            agent.replay_buffer.push((state, next_state, action, reward, done))
    
            state = next_state
            
            # 存入列表
            acc1c__.append(env.state[0])
            acc2c__.append(env.state[1])
            acc3c__.append(env.state[2])
            acc4c__.append(env.state[3])
            #print("acc1c__:",acc1c__)#蓝色
            #print("acc1__:", acc1__)#红色
            # print("acc2c__:",acc2c__)
            # print("acc2__:", acc2__)
            # print("acc3c__:",acc3c__)
            # print("acc3__:", acc3__)
            # print("acc4c__:",acc4c__)
            # print("acc4__:", acc4__)
            action1__.append(action[0])
            action2__.append(action[1])
            action3__.append(action[2])
            action4__.append(action[3])
            
    
            num += 1
            ep_reward += reward
            
            if num >= args.episode_length:#长度>500
                ep_reward__.append(ep_reward)
                S_acc_current1 = 0
                S_acc_current2 = 0
                S_acc_current3 = 0
                S_acc_current4 = 0
                for i in acc1c__:
                    S_acc_current1 += i * i
                for i in acc2c__:
                    S_acc_current2 += i * i
                for i in acc3c__:
                    S_acc_current3 += i * i
                for i in acc4c__:
                    S_acc_current4 += i * i
                S_acc_current1 = S_acc_current1 / len(acc1__)
                S_acc_current2 = S_acc_current2 / len(acc1__)
                S_acc_current3 = S_acc_current3 / len(acc1__)
                S_acc_current4 = S_acc_current4 / len(acc1__)
                S_acc_current = (S_acc_current1 + S_acc_current2 + S_acc_current3 + S_acc_current4)/4
                print("Episode:{} Total Reward: {:.1f} Explore:{:.2f} S_acc_org1:{:.6f} S_acc_org2:{:.6f} S_acc_org3:{:.6f} S_acc_org4:{:.6f} S_acc_org:{:.6f} S_acc_current1:{:.6f} S_acc_current2:{:.6f} S_acc_current3:{:.6f} S_acc_current4:{:.6f} S_acc_current:{:.6f} ".format(
                    i_episode, ep_reward, args.exploration_noise, S_acc_org1,S_acc_org2,S_acc_org3,S_acc_org4,S_acc_org, S_acc_current1,S_acc_current2,S_acc_current3,S_acc_current4,S_acc_current))
                if i_episode % 1 == 0:  # 每4轮输出一次
                    
                    Y1c = fft(acc1c__, nfft)
                    P1c = abs(Y1c/nfft)
                    
                    Y2c = fft(acc2c__, nfft)
                    P2c = abs(Y2c/nfft)
                    
                    Y3c = fft(acc3c__, nfft)
                    P3c = abs(Y3c/nfft)
                    
                    Y4c = fft(acc4c__, nfft)
                    P4c = abs(Y4c/nfft)
                    
                    t = env.tau * np.array(range(0, num))
                    
                    plt.figure(figsize=(16,12))
        
                    plt.subplot(441)
                    plt.plot(t, acc1__, "r--", label="acc1", linewidth=3)
                    plt.plot(t, acc1c__, "b-", label="acc1 Controlled", linewidth=3)
                    plt.xlabel("Time(s)")
                    plt.ylabel("Acc") 
                    plt.legend()
                    
                    plt.subplot(442)
                    plt.plot(t, acc2__, "r--", label="acc2", linewidth=3)
                    plt.plot(t, acc2c__, "b-", label="acc2 Controlled", linewidth=3)
                    plt.ylabel("Acc") 
                    plt.xlabel("Time(s)")
                    plt.legend()
    
                    plt.subplot(443)
                    plt.plot(t, acc3__, "r--", label="acc3", linewidth=3)
                    plt.plot(t, acc3c__, "b-", label="acc3 Controlled", linewidth=3)
                    plt.ylabel("Acc") 
                    plt.xlabel("Time(s)")
                    plt.legend()
                    
                    plt.subplot(444)
                    plt.plot(t, acc4__, "r--", label="acc4", linewidth=3)
                    plt.plot(t, acc4c__, "b-", label="acc4 Controlled", linewidth=3)
                    plt.ylabel("Acc") 
                    plt.xlabel("Time(s)")
                    plt.legend()
                    
                    
                    plt.subplot(445)
                    plt.plot(fre, P1[0:int(nfft/2),],"r--", label="acc1", linewidth=3)
                    plt.plot(fre, P1c[0:int(nfft/2),],"b-", label="acc1 Controlled", linewidth=3)
                    plt.xlabel('Frequency(Hz)')
                    plt.ylabel('FFT')
                    plt.legend()
                    
                    plt.subplot(446)
                    plt.plot(fre, P2[0:int(nfft/2),],"r--", label="acc2", linewidth=3)
                    plt.plot(fre, P2c[0:int(nfft/2),],"b-", label="acc2 Controlled", linewidth=3)
                    plt.xlabel('Frequency(Hz)')
                    plt.ylabel('FFT')
                    plt.legend()
                    
                    plt.subplot(447)
                    plt.plot(fre, P3[0:int(nfft/2),],"r--", label="acc3", linewidth=3)
                    plt.plot(fre, P3c[0:int(nfft/2),],"b-", label="acc3 Controlled", linewidth=3)
                    plt.xlabel('Frequency(Hz)')
                    plt.ylabel('FFT')
                    plt.legend()
                    
                    plt.subplot(448)
                    plt.plot(fre, P4[0:int(nfft/2),],"r--", label="acc4", linewidth=3)
                    plt.plot(fre, P4c[0:int(nfft/2),],"b-", label="acc4 Controlled", linewidth=3)
                    plt.xlabel('Frequency(Hz)')
                    plt.ylabel('FFT')
                    plt.legend()                
                    
                    plt.subplot(449)
                    plt.plot(t, action1__, 'k-', linewidth=3)
                    plt.ylabel("Action 1")
                    plt.xlabel("Time(s)")
                    
                    plt.subplot(4,4,10)
                    plt.plot(t, action2__, 'k-', linewidth=3)
                    plt.ylabel("Action 2")
                    plt.xlabel("Time(s)")
                    
                    plt.subplot(4,4,11)
                    plt.plot(t, action3__, 'k-', linewidth=3)
                    plt.ylabel("Action 3")
                    plt.xlabel("Time(s)")
    
                    plt.subplot(4,4,12)
                    plt.plot(t, action4__, 'k-', linewidth=3)
                    plt.ylabel("Action 4")
                    plt.xlabel("Time(s)")
                    
                    plt.subplot(414)
                    plt.plot(range(len(ep_reward__)), ep_reward__, linewidth=3, color='g')
                    plt.ylabel('Return', size=10)   
                    plt.xlabel('Episode No.', size=10) 
                    
                    # window路径是反斜杠，python不认，需加反斜杠进行转义

                    plt.savefig(path1 + str(i_episode) + '.png')
                    #plt.show()
                    # df_Q = pd.read_excel("Q_data.xlsx")
                    # # print(df_Q)
                    # plt.figure(figsize=(16, 12))
                    # plt.plot(df_Q['current'], "r--", label="current_Q", linewidth=3)
                    # plt.plot(df_Q['target'], "b-", label="target_Q", linewidth=3)
                    # plt.xlabel('times')
                    # plt.ylabel('Q_value')
                    # plt.legend()
                    # plt.savefig(Q_value + str(i_episode) + '.png')
                    #
                    # df_loss = pd.read_excel("loss_data.xlsx")
                    # plt.figure(figsize=(16, 12))
                    # plt.plot(df_loss['loss'], "r--", label="loss", linewidth=3)
                    # plt.xlabel('times')
                    # plt.ylabel('loss_value')
                    # plt.legend()
                    # plt.savefig(loss_value + str(i_episode) + '.png')

                break #结束while循环
        
        #结束while循环，还在for循环内
        if i_episode>=1:  
            agent.update()

        if i_episode % args.save_interval == 0:
            agent.save()
        
    #for循环结束
    reward_list = {"reward":ep_reward__,}
    frame1 = pd.DataFrame(reward_list)
    
    writer = pd.ExcelWriter('reward.xlsx')
    frame1.to_excel(writer, sheet_name='Sheet1', index=False)
    
    writer.save()
    
#%%测试智能体

j = 0
dt = env.tau

f1 = 35
A1 = 2
fai1 = np.pi * 1 / 2

f2 = 40
A2 = 1
fai2 = np.pi * 4 / 3

f3 = 45
A3 = 2
fai3 = np.pi * 5 / 3

acc1__ = []
acc2__ = []
acc3__ = []
acc4__ = []

path = 'motor0on0415_test_true22.xlsx'
df1 = pd.read_excel(path, sheet_name='data', header=None)
x0_array = df1.values[:, 0]

S_acc_org = 0
while j < args.test_num:#300
    # 开启振源, 实际中的振源不可知
    # x0 = A * np.sin(2*np.pi*f * dt * j)
    x0 = x0_array[j]

    # x0 = A1*np.sin(2*np.pi*f1*j*dt + fai1) + A2*np.sin(2*np.pi*f2*j*dt + + fai2) + A3*np.sin(2*np.pi*f3*j*dt + fai3)

    # 电机的动作，神经网络的输出，这里关闭电机
    x1 = 0
    x2 = 0
    x3 = 0
    x4 = 0
    # 状态，实际从传感器读取
    env.state, reward, done = env.step(x0, x1, x2, x3, x4)
    acc1__.append(env.state[0])
    acc2__.append(env.state[1])
    acc3__.append(env.state[2])
    acc4__.append(env.state[3])

    j += 1
S_acc_org1 = 0
S_acc_org2 = 0
S_acc_org3 = 0
S_acc_org4 = 0
for i in acc1__:
    S_acc_org1 += abs(i)
for i in acc2__:
    S_acc_org2 += abs(i)
for i in acc3__:
    S_acc_org3 += abs(i)
for i in acc4__:
    S_acc_org4 += abs(i)
print("acc1__", acc1__)
print("acc2__", acc2__)
print("acc3__", acc3__)
print("acc4__", acc4__)
S_acc_org1 = S_acc_org1 / len(acc1__)
S_acc_org2 = S_acc_org2 / len(acc1__)
S_acc_org3 = S_acc_org3 / len(acc1__)
S_acc_org4 = S_acc_org4 / len(acc1__)
S_acc_org = (S_acc_org1 + S_acc_org2 + S_acc_org3 + S_acc_org4) / 4


Fs = int(1 / env.tau)  # 1/0.001=1000
nfft = 2 * Fs #2000
fre = Fs / nfft * (np.array(range(0, int(nfft/2)))) # 1000/2000*range(0,1000)=0.5*(0,1000)

Y1 = fft(acc1__, nfft)
P1 = abs(Y1 / nfft)

Y2 = fft(acc2__, nfft)
P2 = abs(Y2 / nfft)

Y3 = fft(acc3__, nfft)
P3 = abs(Y3 / nfft)

Y4 = fft(acc4__, nfft)
P4 = abs(Y4 / nfft)

# plt.figure(figsize=(10,6))
# plt.subplot(221)
# plt.plot(fre, P1[0:int(nfft/2),], '-b')
# plt.xlabel('Frequency(Hz)')
# plt.ylabel('Acc1 FFT')
#
# plt.subplot(222)
# plt.plot(fre, P2[0:int(nfft/2),], '-b')
# plt.xlabel('Frequency(Hz)')
# plt.ylabel('Acc2 FFT')
#
# plt.subplot(223)
# plt.plot(fre, P3[0:int(nfft/2),], '-b')
# plt.xlabel('Frequency(Hz)')
# plt.ylabel('Acc3 FFT')
#
# plt.subplot(224)
# plt.plot(fre, P4[0:int(nfft/2),], '-b')
# plt.xlabel('Frequency(Hz)')
# plt.ylabel('Acc4 FFT')

t = env.tau * np.array(range(0, args.test_num))
plt.figure(figsize=(16, 12))

plt.subplot(441)
plt.plot(t, acc1__, "r--", label="acc1", linewidth=3)
plt.xlabel("Time(s)")
plt.ylabel("Acc")
plt.legend()

plt.subplot(442)
plt.plot(t, acc2__, "r--", label="acc2", linewidth=3)
plt.ylabel("Acc")
plt.xlabel("Time(s)")
plt.legend()

plt.subplot(443)
plt.plot(t, acc3__, "r--", label="acc3", linewidth=3)
plt.ylabel("Acc")
plt.xlabel("Time(s)")
plt.legend()

plt.subplot(444)
plt.plot(t, acc4__, "r--", label="acc4", linewidth=3)
plt.ylabel("Acc")
plt.xlabel("Time(s)")
plt.legend()

plt.subplot(445)
plt.plot(fre, P1[0:int(nfft/2), ], "r--", label="acc1", linewidth=3)
plt.xlabel('Frequency(Hz)')
plt.ylabel('FFT')
plt.legend()

plt.subplot(446)
plt.plot(fre, P2[0:int(nfft/2), ], "r--", label="acc2", linewidth=3)
plt.xlabel('Frequency(Hz)')
plt.ylabel('FFT')
plt.legend()

plt.subplot(447)
plt.plot(fre, P3[0:int(nfft/2), ], "r--", label="acc3", linewidth=3)
plt.xlabel('Frequency(Hz)')
plt.ylabel('FFT')
plt.legend()

plt.subplot(448)
plt.plot(fre, P4[0:int(nfft/2), ], "r--", label="acc4", linewidth=3)
plt.xlabel('Frequency(Hz)')
plt.ylabel('FFT')
plt.legend()

# t = env.tau * np.array(range(0, args.episode_length))
# plt.figure(figsize=(16, 12))
# plt.subplot(441)
# plt.plot(t, acc1__, "r--", label="acc1", linewidth=3)
# plt.xlabel("Time(s)")
# plt.ylabel("Acc")
plt.savefig(path + '\\pic\\' + 'org_test.png')


# if args.mode == 'test':
#     print("============ train agent ==============")
#     if args.load_nn == True:
#         agent.load()
#     x0 = -0.0101811886146867
#     env.reset()
#     x1 = [-0.27145207, -0.2609719, -0.25724033, -0.2562204, -0.2562635, -0.25663623, -0.25708833, -0.25764826,
#           -0.2583795, -0.2592984, -0.26035282]
#     x2 = [0.11864785, 0.12713388, 0.13015494, 0.13095517, 0.13091902, 0.1306187, 0.13025519, 0.1298067, 0.1292188,
#           0.12847814, 0.12762754]
#     x3 = [0.09305832, 0.107684866, 0.11287568, 0.11425975, 0.11419414, 0.113675565, 0.11304937, 0.11227633, 0.1112643,
#           0.10998969, 0.10852562]
#     x4 = [0.19572042, 0.21208516, 0.21786208, 0.21941118, 0.2193408, 0.2187626, 0.21806519, 0.21720256, 0.21607421,
#           0.21465284, 0.21301968]
#
#     y1 = [-0.0153030876501895, -0.0140907473595978, -0.013335390776686, -0.01280284482463, -0.0124593449176862,
#           -0.0121850606044709, -0.0117923068309048, -0.0111649895131932, -0.0103395723990611, -0.00942902907559404,
#           -0.00858058173594747]
#     y2 = [0.000319427784639469, 0.00120901591846086, 0.00152894070154164, 0.00162646539765175, 0.00165619385393012,
#           0.00168563388594937, 0.00172754596723008, 0.00175949909516626, 0.00175921287198601, 0.00172886405447697,
#           0.00168706005781824]
#     y3 = [-0.0095475087216081, -0.00849938880640105, -0.0079192279577315, -0.00754998795457392, -0.00732023679893468,
#           -0.00713312230704223, -0.0068665960751641, -0.00645494089674023, -0.00592860252016017, -0.00535789611205152,
#           -0.00482948024752684]
#     y4 = [-0.00543615114394206, -0.00438234263473604, -0.00388722211741298, -0.00362639147240441, -0.00348291426482151,
#           -0.00336630441147938, -0.00319816478851071, -0.00295054952128689, -0.00265175700691504, -0.00234226890906565,
#           -0.00206404143060249]
#
#     for i in range(10):
#
#         env.state, reward, done = env.step(x0, args.action_scale_factor*x1[i], args.action_scale_factor*x2[i], args.action_scale_factor*x3[i], args.action_scale_factor*x4[i])
#         if((abs(env.state[0]-y1[i])>math.pow(10,-7)) | (abs(env.state[1]-y2[i])>math.pow(10,-7)) | (abs(env.state[2]-y3[i])>math.pow(10,-7)) |(abs(env.state[3]-y4[i])>math.pow(10,-7))  ):
#             print("find the error :%d",i)
#             print("%f,%f,%f", env.state[0], y1[i],(abs(env.state[0]-y1[i])>math.pow(10,-7)))
#             print("%f,%f,%f", env.state[1], y2[i],(abs(env.state[1]-y2[i])>math.pow(10,-7)))
#             print("%f,%f,%f", env.state[2], y3[i],(abs(env.state[2]-y3[i])>math.pow(10,-7)))
#             print("%f,%f,%f", env.state[3], y4[i],(abs(env.state[3]-y4[i])>math.pow(10,-7)))
#             break
#     print("全部一样")
if args.mode == 'test':
    print("============ test agent ==============")
    agent.load()
    df1 = pd.read_excel(path, sheet_name='data', header=None)
    x0_array = df1.values[:, 0]

    for i_episode in range(1):#300

        # 进入新的episode之前, 清空列表
        acc1c__ = []
        acc2c__ = []
        acc3c__ = []
        acc4c__ = []

        action1__ = []
        action2__ = []
        action3__ = []
        action4__ = []
        s1_history = 100 * [0.0]
        s2_history = 100 * [0.0]
        s3_history = 100 * [0.0]
        s4_history = 100 * [0.0]

        ep_reward__ = []

        ep_reward = 0
        num = 0
        count = 0
        # 重置
        env.reset()

        s_history = s1_history + s2_history + s3_history + s4_history

        # 获取第一个s
        state = args.state_scale_factor * np.array(s_history)
        while(1):
            action = agent.select_action(state).clip(-1, 1)
            #print(action)
            x0 = x0_array[num]

            env.state, reward, done = env.step(x0, args.action_scale_factor*action[0], args.action_scale_factor*action[1], args.action_scale_factor*action[2], args.action_scale_factor*action[3])

            #s1_history = s1_history * 0.9  报错 TypeError: can't multiply sequence by non-int of type 'float'
            s1_history = [i * 0.5 for i in s1_history]
            s1_history.append(env.state[0])
            s1_history.pop(0)
            s2_history = [i * 0.5 for i in s2_history]
            s2_history.append(env.state[1])
            s2_history.pop(0)
            s3_history = [i * 0.5 for i in s3_history]
            s3_history.append(env.state[2])
            s3_history.pop(0)
            s4_history = [i * 0.5 for i in s4_history]
            s4_history.append(env.state[3])
            s4_history.pop(0)

            s_history = s1_history + s2_history + s3_history + s4_history

            next_state = args.state_scale_factor * np.array(s_history)

            state = next_state

            # 存入列表
            #print(env.state)
            acc1c__.append(env.state[0])
            acc2c__.append(env.state[1])
            acc3c__.append(env.state[2])
            acc4c__.append(env.state[3])

            action1__.append(action[0])
            action2__.append(action[1])
            action3__.append(action[2])
            action4__.append(action[3])

            num += 1
            ep_reward += reward
            print(num)
            if num == args.episode_length:#300
                #print("count",count)

                print(i_episode)
                #if i_episode %0 : #300-1 # 每4轮输出一次
                print("i_episode",i_episode)
                print("args.test_num",args.test_num)
                ep_reward__.append(ep_reward)
                S_acc_current1 = 0
                S_acc_current2 = 0
                S_acc_current3 = 0
                S_acc_current4 = 0
                for i in acc1c__:
                    S_acc_current1 += abs(i)
                for i in acc2c__:
                    S_acc_current2 += abs(i)
                for i in acc3c__:
                    S_acc_current3 += abs(i)
                for i in acc4c__:
                    S_acc_current4 += abs(i)
                S_acc_current1 = S_acc_current1 / len(acc1c__)
                S_acc_current2 = S_acc_current2 / len(acc1c__)
                S_acc_current3 = S_acc_current3 / len(acc1c__)
                S_acc_current4 = S_acc_current4 / len(acc1c__)
                S_acc_current = (S_acc_current1 + S_acc_current2 + S_acc_current3 + S_acc_current4) / 4
                print("acc1cc__", len(acc1c__))
                print(
                    "Episode:{} Total Reward: {:.1f} Explore:{:.2f} S_acc_org1:{:.6f} S_acc_org2:{:.6f} S_acc_org3:{:.6f} S_acc_org4:{:.6f} S_acc_org:{:.6f} S_acc_current1:{:.6f} S_acc_current2:{:.6f} S_acc_current3:{:.6f} S_acc_current4:{:.6f} S_acc_current:{:.6f} ".format(
                        i_episode, ep_reward, args.exploration_noise, S_acc_org1, S_acc_org2, S_acc_org3,
                        S_acc_org4,
                        S_acc_org, S_acc_current1, S_acc_current2, S_acc_current3, S_acc_current4, S_acc_current))


                Y1c = fft(acc1c__, nfft)
                P1c = abs(Y1c/nfft)

                Y2c = fft(acc2c__, nfft)
                P2c = abs(Y2c/nfft)

                Y3c = fft(acc3c__, nfft)
                P3c = abs(Y3c/nfft)

                Y4c = fft(acc4c__, nfft)
                P4c = abs(Y4c/nfft)

                t = env.tau *np.array(range(0, args.test_num))

                plt.figure(figsize=(16,12))

                plt.subplot(341)
                plt.plot(t, acc1__, "r--", label="acc1", linewidth=3)
                plt.plot(t, acc1c__, "b-", label="acc1 Controlled", linewidth=3)
                plt.xlabel("Time(s)")
                plt.ylabel("Acc")
                plt.legend()

                plt.subplot(342)
                plt.plot(t, acc2__, "r--", label="acc2", linewidth=3)
                plt.plot(t, acc2c__, "b-", label="acc2 Controlled", linewidth=3)
                plt.ylabel("Acc")
                plt.xlabel("Time(s)")
                plt.legend()

                plt.subplot(343)
                plt.plot(t, acc3__, "r--", label="acc3", linewidth=3)
                plt.plot(t, acc3c__, "b-", label="acc3 Controlled", linewidth=3)
                plt.ylabel("Acc")
                plt.xlabel("Time(s)")
                plt.legend()

                plt.subplot(344)
                plt.plot(t, acc4__, "r--", label="acc4", linewidth=3)
                plt.plot(t, acc4c__, "b-", label="acc4 Controlled", linewidth=3)
                plt.ylabel("Acc")
                plt.xlabel("Time(s)")
                plt.legend()


                plt.subplot(345)
                plt.plot(fre, P1[0:int(nfft/2),],"r--", label="acc1", linewidth=3)
                plt.plot(fre, P1c[0:int(nfft/2),],"b-", label="acc1 Controlled", linewidth=3)
                plt.xlabel('Frequency(Hz)')
                plt.ylabel('FFT')
                plt.legend()

                plt.subplot(346)
                plt.plot(fre, P2[0:int(nfft/2),],"r--", label="acc2", linewidth=3)
                plt.plot(fre, P2c[0:int(nfft/2),],"b-", label="acc2 Controlled", linewidth=3)
                plt.xlabel('Frequency(Hz)')
                plt.ylabel('FFT')
                plt.legend()

                plt.subplot(347)
                plt.plot(fre, P3[0:int(nfft/2),],"r--", label="acc3", linewidth=3)
                plt.plot(fre, P3c[0:int(nfft/2),],"b-", label="acc3 Controlled", linewidth=3)
                plt.xlabel('Frequency(Hz)')
                plt.ylabel('FFT')
                plt.legend()

                plt.subplot(348)
                plt.plot(fre, P4[0:int(nfft/2),],"r--", label="acc4", linewidth=3)
                plt.plot(fre, P4c[0:int(nfft/2),],"b-", label="acc4 Controlled", linewidth=3)
                plt.xlabel('Frequency(Hz)')
                plt.ylabel('FFT')
                plt.legend()

                plt.subplot(349)
                plt.plot(t, action1__, 'k-', linewidth=3)
                plt.ylabel("Action 1")
                plt.xlabel("Time(s)")

                plt.subplot(3,4,10)
                plt.plot(t, action2__, 'k-', linewidth=3)
                plt.ylabel("Action 2")
                plt.xlabel("Time(s)")

                plt.subplot(3,4,11)
                plt.plot(t, action3__, 'k-', linewidth=3)
                plt.ylabel("Action 3")
                plt.xlabel("Time(s)")

                plt.subplot(3,4,12)
                plt.plot(t, action4__, 'k-', linewidth=3)
                plt.ylabel("Action 4")
                plt.xlabel("Time(s)")


                # window路径是反斜杠，python不认，需加反斜杠进行转义
                plt.savefig(path1 + 'test' + str(i_episode)+'.png')
                break
                #plt.show()

            #break #结束while循环

