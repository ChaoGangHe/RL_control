import argparse
import copy

import os
import numpy as np
import pandas as pd

import torch
import torch.nn as nn
import torch.nn.functional as F
import torch.optim as optim

from tensorboardX import SummaryWriter
from scipy.fftpack import fft
import matplotlib.pyplot as plt
import openpyxl

# %% 超参数
parser = argparse.ArgumentParser()
parser.add_argument('--mode', default='test', type=str)  # mode = 'train' or 'test'
parser.add_argument('--load_nn', default=False, type=bool)  # 是否导入已有网络

parser.add_argument('--tau', default=0.005, type=float)  # target smoothing coefficient

parser.add_argument('--lr_A', default=3e-4, type=float)  # A网络学习率
parser.add_argument('--lr_C', default=3e-4, type=float)  # C网络学习率

parser.add_argument('--gamma', default=0.99, type=int)  # discounted factor
parser.add_argument('--capacity', default=2000, type=int)  # replay buffer size
parser.add_argument('--batch_size', default=128, type=int)  # mini batch size

parser.add_argument('--episode_length', default=300, type=int)  # 回合长度()===x0的长度
parser.add_argument('--save_interval', default=1, type=int)  # 相隔n回合存储一次网络参数
parser.add_argument('--max_episode', default=2001, type=int)  # 回合数
parser.add_argument('--update_iteration', default=300, type=int)  # 每回合更新网络参数的次数

parser.add_argument('--exploration_noise', default=0.2, type=float)  # 探索噪声初值

parser.add_argument('--state_scale_factor', default=0.1, type=float)  # 状态放大系数（放大之后输入网络） 10
parser.add_argument('--action_scale_factor', default=0.02, type=float)  # 动作放大系数（网络输出后放大） 2
parser.add_argument('--test_num', default=300, type=int)  # 训练样本长度
args = parser.parse_args()

device = 'cuda' if torch.cuda.is_available() else 'cpu'
# device = torch.device('cpu')

# %%路径
path = os.getcwd()

path1 = path + '\\pic2\\'
if not os.path.exists('./pic2') == False:
    os.mkdir(path1)

directory = path + '\\nn2\\'
if not os.path.exists('./nn2') == False:
    os.mkdir(directory)

Q_value = path + '\\Q_value\\'
if not os.path.exists('./Q_value') == False:
    os.mkdir(Q_value)

loss_value = path + '\\loss_value\\'
if not os.path.exists('./loss_value') == False:
    os.mkdir(loss_value)

# %%读取传递通道模型参数
df = pd.read_excel('Wts0415.xlsx', sheet_name='Sheet1', header=None)
# df = pd.read_excel('Wts0126.xlsx',sheet_name='Sheet1', header=None)
data = df.values

W01 = data[:, 0]
W02 = data[:, 1]
W03 = data[:, 2]
W04 = data[:, 3]

W11 = data[:, 4]
W12 = data[:, 5]
W13 = data[:, 6]
W14 = data[:, 7]

W21 = data[:, 8]
W22 = data[:, 9]
W23 = data[:, 10]
W24 = data[:, 11]

W31 = data[:, 12]
W32 = data[:, 13]
W33 = data[:, 14]
W34 = data[:, 15]

W41 = data[:, 16]
W42 = data[:, 17]
W43 = data[:, 18]
W44 = data[:, 19]

W01r = W01[::-1]
W02r = W02[::-1]
W03r = W03[::-1]
W04r = W04[::-1]

W11r = W11[::-1]
W12r = W12[::-1]
W13r = W13[::-1]
W14r = W14[::-1]

W21r = W21[::-1]
W22r = W22[::-1]
W23r = W23[::-1]
W24r = W24[::-1]

W31r = W31[::-1]
W32r = W32[::-1]
W33r = W33[::-1]
W34r = W34[::-1]

W41r = W41[::-1]
W42r = W42[::-1]
W43r = W43[::-1]
W44r = W44[::-1]  # (300,1)

# df1 = pd.read_excel('mortor_self.xlsx',sheet_name='motor', header=None)
df1 = pd.read_excel('motor0on0415_test4.xlsx', sheet_name='data', header=None)
# df1 = pd.read_excel('motor0on.xlsx',sheet_name='Data', header=None)
ii = 1
x0_array = df1.values[1 + ii:20001 + ii, 0]

Q_data = openpyxl.Workbook()
sheet_score = Q_data.create_sheet('value', 0)
sheet_score.append(['current', 'target'])
Q_data.save('Q_data.xlsx')

loss_data = openpyxl.Workbook()
sheet_loss = loss_data.create_sheet('value', 0)
sheet_loss.append(['loss', ])
loss_data.save('loss_data.xlsx')
class Env():
    def __init__(self):
        # 振源、电机的输入均是长度300的向量（）
        self.X0 = 300 * [0]
        self.X1 = 300 * [0]
        self.X2 = 300 * [0]
        self.X3 = 300 * [0]
        self.X4 = 300 * [0]
        self.tau = 0.001

    def step(self, x0, x1, x2, x3, x4):
        # x0是振源, x1,x2,x3,x4电机当前输入值、也是Action
        # 更新输入向量
        self.X0.append(x0)
        self.X0.pop(0)

        self.X1.append(x1)
        self.X1.pop(0)

        self.X2.append(x2)
        self.X2.pop(0)

        self.X3.append(x3)
        self.X3.pop(0)

        self.X4.append(x4)
        self.X4.pop(0)

        # 更新传感器的输出
        self.y1 = np.array(self.X0).dot(W01r) + \
                  np.array(self.X1).dot(W11r) + \
                  np.array(self.X2).dot(W21r) + \
                  np.array(self.X3).dot(W31r) + \
                  np.array(self.X4).dot(W41r)

        self.y2 = np.array(self.X0).dot(W02r) + \
                  np.array(self.X1).dot(W12r) + \
                  np.array(self.X2).dot(W22r) + \
                  np.array(self.X3).dot(W32r) + \
                  np.array(self.X4).dot(W42r)

        self.y3 = np.array(self.X0).dot(W03r) + \
                  np.array(self.X1).dot(W13r) + \
                  np.array(self.X2).dot(W23r) + \
                  np.array(self.X3).dot(W33r) + \
                  np.array(self.X4).dot(W43r)

        self.y4 = np.array(self.X0).dot(W04r) + \
                  np.array(self.X1).dot(W14r) + \
                  np.array(self.X2).dot(W24r) + \
                  np.array(self.X3).dot(W34r) + \
                  np.array(self.X4).dot(W44r)

        # self.y1 = self.y1 *0.01
        # self.y2 = self.y2 *0.01
        # self.y3 = self.y3 *0.01
        # self.y4 = self.y4 *0.01
        self.state = np.array([self.y1, self.y2, self.y3, self.y4])

        self.done = 0

        # 定义reward
        self.reward = -100 * (self.y1) ** 2 - 100 * (self.y2) ** 2 - 100 * (self.y3) ** 2 - 100 * (self.y4) ** 2

        return self.state, self.reward, self.done

    def reset(self):
        self.X0 = 300 * [0]
        self.X1 = 300 * [0]
        self.X2 = 300 * [0]
        self.X3 = 300 * [0]
        self.X4 = 300 * [0]
        self.tau = 0.001
        self.state = np.array([0.0, 0.0, 0.0, 0.0])
        return self.state


# %% 定义智能体
class Replay_buffer():
    '''
    Code based on:
    https://github.com/openai/baselines/blob/master/baselines/deepq/replay_buffer.py
    Expects tuples of (state, next_state, action, reward, done)
    '''

    def __init__(self, max_size=args.capacity):
        self.storage = []
        self.max_size = max_size
        self.ptr = 0

    def push(self, data):
        if len(self.storage) == self.max_size:
            self.storage[int(self.ptr)] = data
            self.ptr = (self.ptr + 1) % self.max_size
        else:
            self.storage.append(data)

    def sample(self, batch_size):
        ind = np.random.randint(0, len(self.storage), size=batch_size)
        x, y, u, r, d = [], [], [], [], []

        for i in ind:
            X, Y, U, R, D = self.storage[i]
            x.append(np.array(X, copy=False))
            y.append(np.array(Y, copy=False))
            u.append(np.array(U, copy=False))
            r.append(np.array(R, copy=False))
            d.append(np.array(D, copy=False))

        return np.array(x), np.array(y), np.array(u), np.array(r).reshape(-1, 1), np.array(d).reshape(-1, 1)


class Actor1(nn.Module):
    def __init__(self, state_dim, action_dim, max_action):
        super(Actor1, self).__init__()
        self.l1 = nn.Linear(state_dim, 256)
        self.l2 = nn.Linear(256, 256)
        self.l3 = nn.Linear(256, action_dim)

        self.max_action = max_action

    def forward(self, state):
        a = F.relu(self.l1(state))
        a = F.relu(self.l2(a))
        return self.max_action * torch.tanh(self.l3(a))


class Critic1(nn.Module):
    def __init__(self, state_dim, action_dim):
        super(Critic1, self).__init__()
        # Q1 architecture
        self.l1 = nn.Linear(state_dim + action_dim, 256)
        self.l2 = nn.Linear(256, 256)
        self.l3 = nn.Linear(256, 1)

        # Q2 architecture
        self.l4 = nn.Linear(state_dim + action_dim, 256)
        self.l5 = nn.Linear(256, 256)
        self.l6 = nn.Linear(256, 1)

    def forward(self, state, action):
        sa = torch.cat([state, action], 1)

        q1 = F.relu(self.l1(sa))
        q1 = F.relu(self.l2(q1))
        q1 = self.l3(q1)

        q2 = F.relu(self.l4(sa))
        q2 = F.relu(self.l5(q2))
        q2 = self.l6(q2)
        return q1, q2

    def Q1(self, state, action):
        sa = torch.cat([state, action], 1)

        q1 = F.relu(self.l1(sa))
        q1 = F.relu(self.l2(q1))
        q1 = self.l3(q1)
        return q1


class DDPG(object):
    def __init__(self, state_dim, action_dim, max_action):
        self.actor = Actor1(state_dim, action_dim, max_action).to(device)
        self.actor_target = copy.deepcopy(self.actor)
        self.actor_optimizer = optim.Adam(self.actor.parameters(), lr=args.lr_A)

        self.critic = Critic1(state_dim, action_dim).to(device)
        self.critic_target = copy.deepcopy(self.critic)
        self.critic_optimizer = optim.Adam(self.critic.parameters(), lr=args.lr_C)
        self.replay_buffer = Replay_buffer()
        self.writer = SummaryWriter(directory)

        self.num_critic_update_iteration = 0
        self.num_actor_update_iteration = 0
        self.num_training = 0

    def select_action(self, state):
        state = torch.FloatTensor(state.reshape(1, -1)).to(device)
        return self.actor(state).cpu().data.numpy().flatten()

    def update(self):

        for it in range(args.update_iteration):
            # Sample replay buffer
            x, y, u, r, d = self.replay_buffer.sample(args.batch_size)
            state = torch.FloatTensor(x).to(device)
            action = torch.FloatTensor(u).to(device)
            next_state = torch.FloatTensor(y).to(device)
            done = torch.FloatTensor(1 - d).to(device)
            reward = torch.FloatTensor(r).to(device)

            # Select action according to policy and add clipped noise
            noise = (
                    torch.randn_like(action) * 0.2
            ).clamp(-0.5, 0.5)

            next_action = (
                    self.actor_target(next_state) + noise
            ).clamp(-self.actor.max_action, self.actor.max_action)

            # Compute the target Q value
            target_Q1, target_Q2 = self.critic_target(next_state, next_action)
            target_Q = torch.min(target_Q1, target_Q2)
            target_Q = reward + done * args.gamma * target_Q

            # Get current Q estimates
            current_Q1, current_Q2 = self.critic(state, action)

            # Compute critic loss
            critic_loss = F.mse_loss(current_Q1, target_Q) + F.mse_loss(current_Q2, target_Q)

            # Optimize the critic
            self.critic_optimizer.zero_grad()
            critic_loss.backward()
            self.critic_optimizer.step()

            # Delayed policy updates
            if it % 2 == 0:

                # Compute actor losse
                actor_loss = -self.critic.Q1(state, self.actor(state)).mean()

                # Optimize the actor
                self.actor_optimizer.zero_grad()
                actor_loss.backward()
                self.actor_optimizer.step()

                # Update the frozen target models
                for param, target_param in zip(self.critic.parameters(), self.critic_target.parameters()):
                    target_param.data.copy_(args.tau * param.data + (1 - args.tau) * target_param.data)

                for param, target_param in zip(self.actor.parameters(), self.actor_target.parameters()):
                    target_param.data.copy_(args.tau * param.data + (1 - args.tau) * target_param.data)

            self.num_actor_update_iteration += 1
            self.num_critic_update_iteration += 1

    def save(self):
        torch.save(self.actor.state_dict(), directory + 'actor.pt')
        torch.save(self.critic.state_dict(), directory + 'critic.pt')
        print("====================================")
        print("Model has been saved...")
        print("====================================")

    def load(self):
        self.actor.load_state_dict(torch.load(directory + 'actor.pt'))
        self.critic.load_state_dict(torch.load(directory + 'critic.pt'))
        print("====================================")
        print("model has been loaded...")
        print("====================================")
agent = DDPG(state_dim=400, action_dim=4, max_action=1)
env = Env()
if args.mode == 'train':
    print("============ train agent ==============")
    if args.load_nn == True:
        agent.load()
    x0 = -0.010181189
    x1 = -0.27145207
    x2 = 0.09305832
    x3 = 0.11864785
    x4 = 0.19572042
    env.state, reward, done = env.step(x0, x1, x2, x3, x4)
    print(env.state[0])
    print(env.state[1])
    print(env.state[2])
    print(env.state[3])