import random

import openpyxl
import numpy as np
import pandas as pd

# j = 0
# dt = 0.005
#
# f1 = 35
# A1 = 2
# fai1 = np.pi*1/2
#
# f2 = 40
# A2 = 1
# fai2 = np.pi*4/3
#
# f3 = 45
# A3 = 2
# fai3 = np.pi*5/3
df1 = pd.read_excel('motor0on0415.xlsx',sheet_name='data', header=None)
ii = 1
x0_array = df1.values[:,0]
def generator_excel():
    for excel_num in range(5):
        motor_data = openpyxl.Workbook()
        sheet_motor = motor_data.create_sheet('data', 0)

        for it in range(excel_num*10000,(excel_num+1)*10000):

            # x0 = A1*np.sin(2*np.pi*f1*it*dt + fai1)
            # x1 = A2*np.sin(2*np.pi*f2*it*dt + fai2)
            # x2 = A3*np.sin(2*np.pi*f3*it*dt + fai3)
            sheet_motor.append([x0_array[it]])
            # sheet_loss.append([x1])
            # sheet_loss.append([x2])
        motor_data.save('motor0on0415_test'+str(excel_num)+'.xlsx')
    return
generator_excel()
def generator_excel_test(num,*excel_str):
    count = 0
    for str_i in excel_str:#4

        for times in range(5):
            motor_data = openpyxl.Workbook()
            sheet_motor = motor_data.create_sheet('data', 0)
            df1 = pd.read_excel(str_i, sheet_name='data', header=None)
            x0_array = df1.values[:, 0]
            start = random.randint(1,9600)
            for it in range(num):
                sheet_motor.append([x0_array[it+start]])

            motor_data.save('motor0on0415_test_true' + str(count+8)+str(times) + '.xlsx')
        count = count + 1
str_excel = ["motor0on0415_test0.xlsx","motor0on0415_test1.xlsx","motor0on0415_test2.xlsx","motor0on0415_test3.xlsx"]
generator_excel_test(300,*str_excel)

def generator_excel_test(num,*excel_str):

    for times in range(60):
        motor_data = openpyxl.Workbook()
        sheet_motor = motor_data.create_sheet('data', 0)
        df1 = pd.read_excel(excel_str[0], sheet_name='data', header=None)
        x0_array = df1.values[:, 0]
        start = random.randint(1,50000-300)
        for it in range(num):
            sheet_motor.append([x0_array[it+start]])

        motor_data.save('motor0on0415_' +str(times) + '.xlsx')
str_excel = ["motor0on0415.xlsx"]
generator_excel_test(300,*str_excel)