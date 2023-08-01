# -*- coding: utf-8 -*-
"""
Created on Mon Dec 13 10:41:53 2021

@author: admin

Implementation of Deep Deterministic Policy Gradients (DDPG) with pytorch
riginal paper: https://arxiv.org/abs/1509.02971
Not the author's implementation !

"""
import argparse
import copy
import math
import random
import os

import numpy as np
import pandas as pd
import torch
import torch.nn as nn
import torch.nn.functional as F
import torch.optim as optim

from tensorboardX import SummaryWriter
from scipy.fftpack import fft
import matplotlib.pyplot as plt
import openpyxl

# %% 超参数
parser = argparse.ArgumentParser()
parser.add_argument('--mode', default='train', type=str)  # mode = 'train' or 'test'
parser.add_argument('--load_nn', default=False, type=bool)  # 是否导入已有网络

parser.add_argument('--tau', default=0.002, type=float)  # target smoothing coefficient

parser.add_argument('--lr_A', default=1e-5, type=float)  # A网络学习率
parser.add_argument('--lr_C', default=1e-4, type=float)  # C网络学习率

parser.add_argument('--gamma', default=0.99, type=int)  # discounted factor
parser.add_argument('--capacity', default=2000, type=int)  # replay buffer size
parser.add_argument('--batch_size', default=128, type=int)  # mini batch size

parser.add_argument('--episode_length', default=300, type=int)  # 回合长度()===x0的长度
parser.add_argument('--save_interval', default=1, type=int)  # 相隔n回合存储一次网络参数
parser.add_argument('--max_episode', default=2001, type=int)  # 回合数
parser.add_argument('--update_iteration', default=300, type=int)  # 每回合更新网络参数的次数

parser.add_argument('--exploration_noise', default=0.2, type=float)  # 探索噪声初值

parser.add_argument('--state_scale_factor', default=0.1, type=float)  # 状态放大系数（放大之后输入网络） 10
parser.add_argument('--action_scale_factor', default=0.02, type=float)  # 动作放大系数（网络输出后放大） 2
parser.add_argument('--test_num', default=300, type=int)  # 训练样本长度
args = parser.parse_args()

device = 'cuda' if torch.cuda.is_available() else 'cpu'
# device = torch.device('cpu')

# %%路径
path = os.getcwd()

path1 = path + '\\pic2\\'
if not os.path.exists(path1)==True:
    os.mkdir(path1)

directory = path + '\\nn2\\'
if not os.path.exists(directory)==True:
    os.mkdir(directory)

Q_value = path + '\\Q_value\\'
if not os.path.exists(Q_value)==True:
    os.mkdir(Q_value)

loss_value = path + '\\loss_value\\'
if not os.path.exists(loss_value)==True:
    os.mkdir(loss_value)

# %%读取传递通道模型参数
df = pd.read_excel('Wts0415.xlsx', sheet_name='Sheet1', header=None)
#df = pd.read_excel('Wts0101.xlsx',sheet_name='Sheet1', header=None)
data = df.values

W01 = data[:, 0]
W02 = data[:, 1]
W03 = data[:, 2]
W04 = data[:, 3]

W11 = data[:, 4]
W12 = data[:, 5]
W13 = data[:, 6]
W14 = data[:, 7]

W21 = data[:, 8]
W22 = data[:, 9]
W23 = data[:, 10]
W24 = data[:, 11]

W31 = data[:, 12]
W32 = data[:, 13]
W33 = data[:, 14]
W34 = data[:, 15]

W41 = data[:, 16]
W42 = data[:, 17]
W43 = data[:, 18]
W44 = data[:, 19]

W01r = W01[::-1]
W02r = W02[::-1]
W03r = W03[::-1]
W04r = W04[::-1]

W11r = W11[::-1]
W12r = W12[::-1]
W13r = W13[::-1]
W14r = W14[::-1]

W21r = W21[::-1]
W22r = W22[::-1]
W23r = W23[::-1]
W24r = W24[::-1]

W31r = W31[::-1]
W32r = W32[::-1]
W33r = W33[::-1]
W34r = W34[::-1]

W41r = W41[::-1]
W42r = W42[::-1]
W43r = W43[::-1]
W44r = W44[::-1]  # (300,1)

Q_data = openpyxl.Workbook()
sheet_score = Q_data.create_sheet('value', 0)
sheet_score.append(['current', 'target'])
Q_data.save('Q_data.xlsx')

loss_data = openpyxl.Workbook()
sheet_loss = loss_data.create_sheet('value', 0)
sheet_loss.append(['loss', ])
loss_data.save('loss_data.xlsx')


# %% 定义环境
class Env():
    def __init__(self):
        # 振源、电机的输入均是长度300的向量（）
        self.X0 = 300 * [0]
        self.X1 = 300 * [0]
        self.X2 = 300 * [0]
        self.X3 = 300 * [0]
        self.X4 = 300 * [0]
        self.tau = 0.001

    def step(self, x0, x1, x2, x3, x4):
        # x0是振源, x1,x2,x3,x4电机当前输入值、也是Action
        # 更新输入向量
        self.X0.append(x0)
        self.X0.pop(0)

        self.X1.append(x1)
        self.X1.pop(0)

        self.X2.append(x2)
        self.X2.pop(0)

        self.X3.append(x3)
        self.X3.pop(0)

        self.X4.append(x4)
        self.X4.pop(0)

        # 更新传感器的输出
        self.y1 = np.array(self.X0).dot(W01r) + \
                  np.array(self.X1).dot(W11r) + \
                  np.array(self.X2).dot(W21r) + \
                  np.array(self.X3).dot(W31r) + \
                  np.array(self.X4).dot(W41r)

        self.y2 = np.array(self.X0).dot(W02r) + \
                  np.array(self.X1).dot(W12r) + \
                  np.array(self.X2).dot(W22r) + \
                  np.array(self.X3).dot(W32r) + \
                  np.array(self.X4).dot(W42r)

        self.y3 = np.array(self.X0).dot(W03r) + \
                  np.array(self.X1).dot(W13r) + \
                  np.array(self.X2).dot(W23r) + \
                  np.array(self.X3).dot(W33r) + \
                  np.array(self.X4).dot(W43r)

        self.y4 = np.array(self.X0).dot(W04r) + \
                  np.array(self.X1).dot(W14r) + \
                  np.array(self.X2).dot(W24r) + \
                  np.array(self.X3).dot(W34r) + \
                  np.array(self.X4).dot(W44r)

        self.y1 = self.y1 *50
        self.y2 = self.y2 *50
        self.y3 = self.y3 *50
        self.y4 = self.y4 *50
        self.state = np.array([self.y1, self.y2, self.y3, self.y4])

        self.done = 0

        # 定义reward
        self.reward = -(self.y1) ** 2 - (self.y2) ** 2 -  (self.y3) ** 2 -  (self.y4) ** 2

        return self.state, self.reward, self.done

    def reset(self):
        self.X0 = 300 * [0]
        self.X1 = 300 * [0]
        self.X2 = 300 * [0]
        self.X3 = 300 * [0]
        self.X4 = 300 * [0]
        self.tau = 0.001
        self.state = np.array([0.0, 0.0, 0.0, 0.0])
        return self.state


# %% 定义智能体
class Replay_buffer():
    '''
    Code based on:
    https://github.com/openai/baselines/blob/master/baselines/deepq/replay_buffer.py
    Expects tuples of (state, next_state, action, reward, done)
    '''

    def __init__(self, max_size=args.capacity):
        self.storage = []
        self.max_size = max_size
        self.ptr = 0

    def push(self, data):
        if len(self.storage) == self.max_size:
            self.storage[int(self.ptr)] = data
            self.ptr = (self.ptr + 1) % self.max_size
        else:
            self.storage.append(data)

    def sample(self, batch_size):
        ind = np.random.randint(0, len(self.storage), size=batch_size)
        x, y, u, r, d = [], [], [], [], []

        for i in ind:
            X, Y, U, R, D = self.storage[i]
            x.append(np.array(X, copy=False))
            y.append(np.array(Y, copy=False))
            u.append(np.array(U, copy=False))
            r.append(np.array(R, copy=False))
            d.append(np.array(D, copy=False))

        return np.array(x), np.array(y), np.array(u), np.array(r).reshape(-1, 1), np.array(d).reshape(-1, 1)

class Actor1(nn.Module):

    def __init__(self, state_dim, action_dim, max_action):
        super(Actor1, self).__init__()

        self.conv = nn.Conv1d(in_channels=1,out_channels=1,kernel_size=3)
        #self.ReLU = nn.Tanh()
        self.MaxPool1d = nn.MaxPool1d(kernel_size=2,stride=1)

        self.l1 = nn.Linear(97,50)
        self.l2 = nn.Linear(50,20)
        self.l3 = nn.Linear(77,40)
        self.l4 = nn.Linear(40,20)
        self.l5 = nn.Linear(20, action_dim)
        self.conv1 = nn.Conv1d(in_channels=1, out_channels=1, kernel_size=3)
        self.MaxPool1d_2 = nn.MaxPool1d(kernel_size=2, stride=1)
        self.max_action = max_action

    def forward(self, state):
        in1, in2, in3, in4 = state.chunk(4, 1)
        c1 = torch.tanh(self.conv(in1.view(in1.shape[0],1,100)))
        c2 = torch.tanh(self.conv(in2.view(in1.shape[0],1,100)))
        c3 = torch.tanh(self.conv(in3.view(in1.shape[0],1,100)))
        c4 = torch.tanh(self.conv(in4.view(in1.shape[0],1,100)))#torch.Size([1, 1, 98])

        c1 = torch.tanh(self.MaxPool1d(c1))
        c2 = torch.tanh(self.MaxPool1d(c2))
        c3 = torch.tanh(self.MaxPool1d(c3))
        c4 = torch.tanh(self.MaxPool1d(c4))

        c1 = c1.view(c1.shape[0], c1.shape[2])
        c2 = c2.view(c2.shape[0], c2.shape[2])
        c3 = c3.view(c3.shape[0], c3.shape[2])
        c4 = c4.view(c4.shape[0], c4.shape[2])

        c1 = torch.tanh(self.l1(c1))
        c2 = torch.tanh(self.l1(c2))#torch.Size([1, 1, 98]) --->#torch.Size([1, 1, 50])
        c3 = torch.tanh(self.l1(c3))
        c4 = torch.tanh(self.l1(c4))

        c1 = torch.tanh(self.l2(c1))
        c2 = torch.tanh(self.l2(c2))#torch.Size([1, 1, 98]) --->#torch.Size([1, 1, 20])
        c3 = torch.tanh(self.l2(c3))
        c4 = torch.tanh(self.l2(c4))

        #x = torch.cat((c1,c2,c3,c4),2)  #torch.Size([1, 1, 20]) --->#torch.Size([1, 1, 80])
        x = torch.cat((c1, c2, c3, c4), 1)
        #print(x.shape)
        x = torch.tanh(self.conv1(x.view(x.shape[0],1,80)))
        x = torch.tanh(self.MaxPool1d(x))
        x = x.view(x.shape[0], x.shape[2])
        x = torch.tanh(self.l3(x))
        x = torch.tanh(self.l4(x))
        #x = self.MaxPool1d(x)
        #print(x.shape)
        #x = x.view(x.shape[0],x.shape[2])               #torch.Size([1, 386])
        x = torch.tanh(self.l5(x))
        return self.max_action * x

class Critic1(nn.Module):
    def __init__(self, state_dim, action_dim):
        super(Critic1, self).__init__()
        # Q1 architecture

        self.conv_Q1 = nn.Conv1d(in_channels=1, out_channels=1, kernel_size=3)
        self.conv_Q1_2 = nn.Conv1d(in_channels=1, out_channels=1, kernel_size=3)
        self.MaxPool1d_1 = nn.MaxPool1d(kernel_size=2, stride=1)
        self.MaxPool1d_2 = nn.MaxPool1d(kernel_size=2, stride=1)
        self.ls1 = nn.Linear(97, 100)
        self.ls2 = nn.Linear(100, 40)

        self.la1 = nn.Linear(action_dim, 40)
        self.la2 = nn.Linear(40, 40)

        self.l3 = nn.Linear(197, 100)
        self.l4 = nn.Linear(100, 20)
        self.l5 = nn.Linear(20, 1)


        self.conv_Q2 = nn.Conv1d(in_channels=1, out_channels=1, kernel_size=3)
        self.conv_Q2_2 = nn.Conv1d(in_channels=1, out_channels=1, kernel_size=3)
        self.MaxPool1d_Q2_1 = nn.MaxPool1d(kernel_size=2, stride=1)
        self.MaxPool1d_Q2_2 = nn.MaxPool1d(kernel_size=2, stride=1)
        self.ls1_Q2 = nn.Linear(97, 100)
        self.ls2_Q2 = nn.Linear(100, 40)

        self.la1_Q2 = nn.Linear(action_dim, 40)
        self.la2_Q2 = nn.Linear(40, 40)

        self.l3_Q2 = nn.Linear(197, 100)
        self.l4_Q2 = nn.Linear(100, 20)
        self.l5_Q2 = nn.Linear(20, 1)


    def forward(self, state, action):
        in1, in2, in3, in4 = state.chunk(4, 1)  # torch.Size([1, 100])
        c1 = torch.relu(self.conv_Q1(in1.view(in1.shape[0], 1, 100)))  # torch.Size([1, 1, 100]) -->torch.Size([1, 1, 98])
        c2 = torch.relu(self.conv_Q1(in2.view(in1.shape[0], 1, 100)))
        c3 = torch.relu(self.conv_Q1(in3.view(in1.shape[0], 1, 100)))
        c4 = torch.relu(self.conv_Q1(in4.view(in1.shape[0], 1, 100)))

        c1 = torch.relu(self.MaxPool1d_1(c1))
        c2 = torch.relu(self.MaxPool1d_1(c2))
        c3 = torch.relu(self.MaxPool1d_1(c3))
        c4 = torch.relu(self.MaxPool1d_1(c4))

        c1 = c1.view(c1.shape[0], c1.shape[2])
        c2 = c2.view(c2.shape[0], c2.shape[2])# torch.Size([1, 1, 98]) -->torch.Size([1, 98])
        c3 = c3.view(c3.shape[0], c3.shape[2])
        c4 = c4.view(c4.shape[0], c4.shape[2])

        c1 = torch.relu(self.ls1(c1))
        c2 = torch.relu(self.ls1(c2))##torch.Size([1, 98])->[1,100]
        c3 = torch.relu(self.ls1(c3))
        c4 = torch.relu(self.ls1(c4))

        c1 = torch.relu(self.ls2(c1))
        c2 = torch.relu(self.ls2(c2))##torch.Size([1, 100])->[1,40]
        c3 = torch.relu(self.ls2(c3))
        c4 = torch.relu(self.ls2(c4))

        a = torch.relu(self.la1(action))##torch.Size([1, 4])->[1,40]
        a = torch.relu(self.la2(a))##torch.Size([1, 40])->[1,40]

        #a = action.view(action.shape[0],1,4)
        q1 = torch.cat((c1, c2, c3, c4,a), 1)  # torch.Size([1, 396])
        q1 = torch.relu(self.conv_Q1_2(q1.view(q1.shape[0],1,200)))
        q1 = torch.relu(self.MaxPool1d_2(q1))

        q1 = q1.view(q1.shape[0], q1.shape[2])
        q1 = torch.relu(self.l3(q1))
        q1 = torch.relu(self.l4(q1))
        q1 = self.l5(q1)



        in1, in2, in3, in4 = state.chunk(4, 1)  # torch.Size([1, 100])
        c1 = torch.relu(self.conv_Q2(in1.view(in1.shape[0], 1, 100)))  # torch.Size([1, 1, 100]) -->torch.Size([1, 1, 98])
        c2 = torch.relu(self.conv_Q2(in2.view(in1.shape[0], 1, 100)))
        c3 = torch.relu(self.conv_Q2(in3.view(in1.shape[0], 1, 100)))
        c4 = torch.relu(self.conv_Q2(in4.view(in1.shape[0], 1, 100)))
        c1 = torch.relu(self.MaxPool1d_Q2_1(c1))
        c2 = torch.relu(self.MaxPool1d_Q2_1(c2))
        c3 = torch.relu(self.MaxPool1d_Q2_1(c3))
        c4 = torch.relu(self.MaxPool1d_Q2_1(c4))

        c1 = c1.view(c1.shape[0], c1.shape[2])
        c2 = c2.view(c2.shape[0], c2.shape[2])  # torch.Size([1, 1, 98]) -->torch.Size([1, 98])
        c3 = c3.view(c3.shape[0], c3.shape[2])
        c4 = c4.view(c4.shape[0], c4.shape[2])

        c1 = torch.relu(self.ls1_Q2(c1))
        c2 = torch.relu(self.ls1_Q2(c2))  ##torch.Size([1, 98])->[1,100]
        c3 = torch.relu(self.ls1_Q2(c3))
        c4 = torch.relu(self.ls1_Q2(c4))

        c1 = torch.relu(self.ls2_Q2(c1))
        c2 = torch.relu(self.ls2_Q2(c2))  ##torch.Size([1, 100])->[1,40]
        c3 = torch.relu(self.ls2_Q2(c3))
        c4 = torch.relu(self.ls2_Q2(c4))

        a = torch.relu(self.la1_Q2(action))  ##torch.Size([1, 4])->[1,40]
        a = torch.relu(self.la2_Q2(a))  ##torch.Size([1, 40])->[1,40]

        q2 = torch.cat((c1, c2, c3, c4, a), 1)  # torch.Size([1, 396])
        q2 = torch.relu(self.conv_Q2_2(q2.view(q2.shape[0],1,200)))
        q2 = torch.relu(self.MaxPool1d_Q2_2(q2))
        q2 = q2.view(q2.shape[0], q2.shape[2])
        q2 = torch.relu(self.l3(q2))
        q2 = torch.relu(self.l4(q2))
        q2 = self.l5(q2)
        return q1,q2

    def Q1(self, state, action):
        in1, in2, in3, in4 = state.chunk(4, 1)  # torch.Size([1, 100])
        c1 = torch.relu(
            self.conv_Q1(in1.view(in1.shape[0], 1, 100)))  # torch.Size([1, 1, 100]) -->torch.Size([1, 1, 98])
        c2 = torch.relu(self.conv_Q1(in2.view(in1.shape[0], 1, 100)))
        c3 = torch.relu(self.conv_Q1(in3.view(in1.shape[0], 1, 100)))
        c4 = torch.relu(self.conv_Q1(in4.view(in1.shape[0], 1, 100)))

        c1 = torch.relu(self.MaxPool1d_1(c1))
        c2 = torch.relu(self.MaxPool1d_1(c2))
        c3 = torch.relu(self.MaxPool1d_1(c3))
        c4 = torch.relu(self.MaxPool1d_1(c4))

        c1 = c1.view(c1.shape[0], c1.shape[2])
        c2 = c2.view(c2.shape[0], c2.shape[2])  # torch.Size([1, 1, 98]) -->torch.Size([1, 98])
        c3 = c3.view(c3.shape[0], c3.shape[2])
        c4 = c4.view(c4.shape[0], c4.shape[2])

        c1 = torch.relu(self.ls1(c1))
        c2 = torch.relu(self.ls1(c2))  ##torch.Size([1, 98])->[1,100]
        c3 = torch.relu(self.ls1(c3))
        c4 = torch.relu(self.ls1(c4))

        c1 = torch.relu(self.ls2(c1))
        c2 = torch.relu(self.ls2(c2))  ##torch.Size([1, 100])->[1,40]
        c3 = torch.relu(self.ls2(c3))
        c4 = torch.relu(self.ls2(c4))

        a = torch.relu(self.la1(action))  ##torch.Size([1, 4])->[1,40]
        a = torch.relu(self.la2(a))  ##torch.Size([1, 40])->[1,40]

        # a = action.view(action.shape[0],1,4)
        q1 = torch.cat((c1, c2, c3, c4, a), 1)  # torch.Size([1, 396])
        q1 = torch.relu(self.conv_Q1_2(q1.view(q1.shape[0], 1, 200)))
        q1 = torch.relu(self.MaxPool1d_2(q1))

        q1 = q1.view(q1.shape[0], q1.shape[2])
        q1 = torch.relu(self.l3(q1))
        q1 = torch.relu(self.l4(q1))
        q1 = self.l5(q1)
        return q1


class DDPG(object):
    def __init__(self, state_dim, action_dim, max_action):
        self.actor = Actor1(state_dim, action_dim, max_action).to(device)
        self.actor_target = copy.deepcopy(self.actor)
        self.actor_optimizer = optim.Adam(self.actor.parameters(), lr=args.lr_A)

        self.critic = Critic1(state_dim, action_dim).to(device)
        self.critic_target = copy.deepcopy(self.critic)
        self.critic_optimizer = optim.Adam(self.critic.parameters(), lr=args.lr_C)
        self.replay_buffer = Replay_buffer()
        self.writer = SummaryWriter(directory)

        self.num_critic_update_iteration = 0
        self.num_actor_update_iteration = 0
        self.num_training = 0
        self.episode = 0
        self.reward = -100000
        self.i_episode = 0

    def select_action(self, state):
        state = torch.FloatTensor(state.reshape(1, -1)).to(device)
        #print(state.shape)
        return self.actor(state).cpu().data.numpy().flatten()

    def update(self):

        for it in range(args.update_iteration):
            # Sample replay buffer
            x, y, u, r, d = self.replay_buffer.sample(args.batch_size)
            state = torch.FloatTensor(x).to(device)
            action = torch.FloatTensor(u).to(device)
            next_state = torch.FloatTensor(y).to(device)
            done = torch.FloatTensor(1 - d).to(device)
            reward = torch.FloatTensor(r).to(device)

            # Select action according to policy and add clipped noise
            noise = (
                    torch.randn_like(action) * 0.2
            ).clamp(-0.5, 0.5)

            next_action = (
                    self.actor_target(next_state) + noise
            ).clamp(-self.actor.max_action, self.actor.max_action)

            # Compute the target Q value
            target_Q1, target_Q2 = self.critic_target(next_state, next_action)
            target_Q = torch.min(target_Q1, target_Q2)
            target_Q = reward + done * args.gamma * target_Q

            # Get current Q estimates
            current_Q1, current_Q2 = self.critic(state, action)

            # Compute critic loss
            critic_loss = F.mse_loss(current_Q1, target_Q) + F.mse_loss(current_Q2, target_Q)

            # Optimize the critic
            self.critic_optimizer.zero_grad()
            critic_loss.backward()
            self.critic_optimizer.step()

            # Delayed policy updates
            if it % 2 == 0:

                # Compute actor losse
                actor_loss = -self.critic.Q1(state, self.actor(state)).mean()

                # Optimize the actor
                self.actor_optimizer.zero_grad()
                actor_loss.backward()
                self.actor_optimizer.step()

                # Update the frozen target models
                for param, target_param in zip(self.critic.parameters(), self.critic_target.parameters()):
                    target_param.data.copy_(args.tau * param.data + (1 - args.tau) * target_param.data)

                for param, target_param in zip(self.actor.parameters(), self.actor_target.parameters()):
                    target_param.data.copy_(args.tau * param.data + (1 - args.tau) * target_param.data)

            self.num_actor_update_iteration += 1
            self.num_critic_update_iteration += 1

    def save(self):
        torch.save(self.actor.state_dict(), directory + 'actor.pt')
        torch.save(self.critic.state_dict(), directory + 'critic.pt')
        print("====================================")
        print("Model has been saved...")
        print("====================================")

    def load(self):
        self.actor.load_state_dict(torch.load(directory + 'actor.pt'))
        self.critic.load_state_dict(torch.load(directory + 'critic.pt'))
        print("====================================")
        print("model has been loaded...")
        print("====================================")





#df1 = pd.read_excel('motor0on0101.xlsx', sheet_name='data', header=None)
df1 = pd.read_excel('motor0on0415_test4.xlsx', sheet_name='data', header=None)
agent = DDPG(state_dim=400, action_dim=4, max_action=1)
env_org = Env()
env = Env()
# %%训练智能体
ii = 1
x0_array = df1.values[1 + ii:20001 + ii, 0]
if args.mode == 'train':
    print("============ train agent ==============")
    if args.load_nn == True:
        agent.load()
    S_acc_current = 0
    ep_reward__ = []


    for i_episode in range(args.max_episode):  # 2000

        env.reset()
        env_org.reset()

        #每一个episode随机截取
        # start= random.randint(2, 10000 - args.episode_length)
        # #start = random.randint(2, 30000 - args.episode_length)
        # x0_array = df1.values[start:args.episode_length + start, 0]

        # %% 测试环境
        acc1__ = []
        acc2__ = []
        acc3__ = []
        acc4__ = []

        j = 0
        S_acc_org = 0
        while j <= args.episode_length - 1:
            x0 = x0_array[j]
            # 电机的动作，神经网络的输出，这里关闭电机
            x1 = 0
            x2 = 0
            x3 = 0
            x4 = 0
            # 状态，实际从传感器读取
            env.state, reward, done = env.step(x0, x1, x2, x3, x4)
            acc1__.append(env.state[0])
            acc2__.append(env.state[1])
            acc3__.append(env.state[2])
            acc4__.append(env.state[3])
            j += 1

        S_acc_org1 = 0
        S_acc_org2 = 0
        S_acc_org3 = 0
        S_acc_org4 = 0
        for i in acc1__:
            S_acc_org1 += i * i
        for i in acc2__:
            S_acc_org2 += i * i
        for i in acc3__:
            S_acc_org3 += i * i
        for i in acc4__:
            S_acc_org4 += i * i
        S_acc_org1 = S_acc_org1 / len(acc1__)
        S_acc_org2 = S_acc_org2 / len(acc1__)
        S_acc_org3 = S_acc_org3 / len(acc1__)
        S_acc_org4 = S_acc_org4 / len(acc1__)
        S_acc_org = (S_acc_org1 + S_acc_org2 + S_acc_org3 + S_acc_org4) / 4
        Fs = int(1 / env.tau)
        nfft = 2 * Fs
        fre = Fs / nfft * (np.array(range(0, int(nfft / 2))))

        Y1 = fft(acc1__, nfft)
        P1 = abs(Y1 / nfft)

        Y2 = fft(acc2__, nfft)
        P2 = abs(Y2 / nfft)

        Y3 = fft(acc3__, nfft)
        P3 = abs(Y3 / nfft)

        Y4 = fft(acc4__, nfft)
        P4 = abs(Y4 / nfft)

        # 进入新的episode之前, 清空列表
        acc1c__ = []
        acc2c__ = []
        acc3c__ = []
        acc4c__ = []

        action1__ = []
        action2__ = []
        action3__ = []
        action4__ = []

        ep_reward = 0
        num = 0  # (0,300)

        # 重置


        s1_history = 100 * [0.0]
        s2_history = 100 * [0.0]
        s3_history = 100 * [0.0]
        s4_history = 100 * [0.0]
        s_history = s1_history + s2_history + s3_history + s4_history

        # 获取第一个s
        state = args.state_scale_factor * np.array(s_history)

        args.exploration_noise *= 0.999  # 每回合噪声衰减

        while (1):
            action = agent.select_action(state)
            # action[2] = 0
            # action[3] = 0
            action = (action + np.random.normal(0, args.exploration_noise, size=4)).clip(-1, 1)

            # x0 = A*np.sin(2*np.pi*f*num*dt)
            x0 = x0_array[num]

            env.state, reward, done = env.step(x0, args.action_scale_factor * action[0],
                                               args.action_scale_factor * action[1],
                                               args.action_scale_factor * action[2],
                                               args.action_scale_factor * action[3])

            s1_history = [i * 0.5 for i in s1_history]
            s1_history.append(env.state[0])
            s1_history.pop(0)
            s2_history = [i * 0.5 for i in s2_history]
            s2_history.append(env.state[1])
            s2_history.pop(0)
            s3_history = [i * 0.5 for i in s3_history]
            s3_history.append(env.state[2])
            s3_history.pop(0)
            s4_history = [i * 0.5 for i in s4_history]
            s4_history.append(env.state[3])
            s4_history.pop(0)

            s_history = s1_history + s2_history + s3_history + s4_history

            next_state = args.state_scale_factor * np.array(s_history)

            agent.replay_buffer.push((state, next_state, action, reward, done))

            state = next_state

            # 存入列表
            acc1c__.append(env.state[0])
            acc2c__.append(env.state[1])
            acc3c__.append(env.state[2])
            acc4c__.append(env.state[3])
            action1__.append(action[0])
            action2__.append(action[1])
            action3__.append(action[2])
            action4__.append(action[3])

            num += 1
            ep_reward += reward

            if num >= args.episode_length:  # 长度>300
                ep_reward__.append(ep_reward)
                S_acc_current1 = 0
                S_acc_current2 = 0
                S_acc_current3 = 0
                S_acc_current4 = 0
                for i in acc1c__:
                    S_acc_current1 += i * i
                for i in acc2c__:
                    S_acc_current2 += i * i
                for i in acc3c__:
                    S_acc_current3 += i * i
                for i in acc4c__:
                    S_acc_current4 += i * i
                S_acc_current1 = S_acc_current1 / len(acc1__)
                S_acc_current2 = S_acc_current2 / len(acc1__)
                S_acc_current3 = S_acc_current3 / len(acc1__)
                S_acc_current4 = S_acc_current4 / len(acc1__)
                S_acc_current = (S_acc_current1 + S_acc_current2 + S_acc_current3 + S_acc_current4) / 4
                print(
                    "Episode:{} Total Reward: {:.1f} Explore:{:.2f} S_acc_org1:{:.6f} S_acc_org2:{:.6f} S_acc_org3:{:.6f} S_acc_org4:{:.6f} S_acc_org:{:.6f} S_acc_current1:{:.6f} S_acc_current2:{:.6f} S_acc_current3:{:.6f} S_acc_current4:{:.6f} S_acc_current:{:.6f} ".format(
                        i_episode, ep_reward, args.exploration_noise, S_acc_org1, S_acc_org2, S_acc_org3, S_acc_org4,
                        S_acc_org, S_acc_current1, S_acc_current2, S_acc_current3, S_acc_current4, S_acc_current))
                #if i_episode % 2 == 0:  # 每4轮输出一次

                Y1c = fft(acc1c__, nfft)
                P1c = abs(Y1c / nfft)

                Y2c = fft(acc2c__, nfft)
                P2c = abs(Y2c / nfft)

                Y3c = fft(acc3c__, nfft)
                P3c = abs(Y3c / nfft)

                Y4c = fft(acc4c__, nfft)
                P4c = abs(Y4c / nfft)

                t = env.tau * np.array(range(0, num))

                plt.figure(figsize=(16, 12))

                plt.subplot(441)
                plt.plot(t, acc1__, "r--", label="acc1", linewidth=3)
                plt.plot(t, acc1c__, "b-", label="acc1 Controlled", linewidth=3)
                plt.xlabel("Time(s)")
                plt.ylabel("Acc")
                plt.legend()

                plt.subplot(442)
                plt.plot(t, acc2__, "r--", label="acc2", linewidth=3)
                plt.plot(t, acc2c__, "b-", label="acc2 Controlled", linewidth=3)
                plt.ylabel("Acc")
                plt.xlabel("Time(s)")
                plt.legend()

                plt.subplot(443)
                plt.plot(t, acc3__, "r--", label="acc3", linewidth=3)
                plt.plot(t, acc3c__, "b-", label="acc3 Controlled", linewidth=3)
                plt.ylabel("Acc")
                plt.xlabel("Time(s)")
                plt.legend()

                plt.subplot(444)
                plt.plot(t, acc4__, "r--", label="acc4", linewidth=3)
                plt.plot(t, acc4c__, "b-", label="acc4 Controlled", linewidth=3)
                plt.ylabel("Acc")
                plt.xlabel("Time(s)")
                plt.legend()

                plt.subplot(445)
                plt.plot(fre, P1[0:int(nfft / 2), ], "r--", label="acc1", linewidth=3)
                plt.plot(fre, P1c[0:int(nfft / 2), ], "b-", label="acc1 Controlled", linewidth=3)
                plt.xlabel('Frequency(Hz)')
                plt.ylabel('FFT')
                plt.legend()

                plt.subplot(446)
                plt.plot(fre, P2[0:int(nfft / 2), ], "r--", label="acc2", linewidth=3)
                plt.plot(fre, P2c[0:int(nfft / 2), ], "b-", label="acc2 Controlled", linewidth=3)
                plt.xlabel('Frequency(Hz)')
                plt.ylabel('FFT')
                plt.legend()

                plt.subplot(447)
                plt.plot(fre, P3[0:int(nfft / 2), ], "r--", label="acc3", linewidth=3)
                plt.plot(fre, P3c[0:int(nfft / 2), ], "b-", label="acc3 Controlled", linewidth=3)
                plt.xlabel('Frequency(Hz)')
                plt.ylabel('FFT')
                plt.legend()

                plt.subplot(448)
                plt.plot(fre, P4[0:int(nfft / 2), ], "r--", label="acc4", linewidth=3)
                plt.plot(fre, P4c[0:int(nfft / 2), ], "b-", label="acc4 Controlled", linewidth=3)
                plt.xlabel('Frequency(Hz)')
                plt.ylabel('FFT')
                plt.legend()

                plt.subplot(449)
                plt.plot(t, action1__, 'k-', linewidth=3)
                plt.ylabel("Action 1")
                plt.xlabel("Time(s)")

                plt.subplot(4, 4, 10)
                plt.plot(t, action2__, 'k-', linewidth=3)
                plt.ylabel("Action 2")
                plt.xlabel("Time(s)")

                plt.subplot(4, 4, 11)
                plt.plot(t, action3__, 'k-', linewidth=3)
                plt.ylabel("Action 3")
                plt.xlabel("Time(s)")

                plt.subplot(4, 4, 12)
                plt.plot(t, action4__, 'k-', linewidth=3)
                plt.ylabel("Action 4")
                plt.xlabel("Time(s)")

                plt.subplot(414)
                plt.plot(range(len(ep_reward__)), ep_reward__, linewidth=3, color='g')
                plt.ylabel('Return', size=10)
                plt.xlabel('Episode No.', size=10)

                # window路径是反斜杠，python不认，需加反斜杠进行转义
                plt.savefig(path1 + str(i_episode) + '.png')


                break  # 结束while循环

        # 结束while循环，还在for循环内
        if i_episode >= 1:
            agent.update()

        if i_episode % args.save_interval == 0:
            # 保存模型
            if ep_reward > agent.reward:
                agent.reward = ep_reward
                agent.i_episode = i_episode
                agent.save()
            else:
                print("Model is worse,not save ,reward:", agent.reward, "i_episode:", agent.i_episode)

    # for循环结束
    reward_list = {"reward": ep_reward__, }
    frame1 = pd.DataFrame(reward_list)

    writer = pd.ExcelWriter('reward.xlsx')
    frame1.to_excel(writer, sheet_name='Sheet1', index=False)

    writer.save()




